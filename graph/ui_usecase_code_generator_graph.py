# @Time    : 2026/3/2 15:47
# @Author  : Yun
# @FileName: ui_use_case_code_generator_graph
# @Software: PyCharm
# @Desc    :
from langchain.chat_models import init_chat_model
from langchain_core.language_models import BaseChatModel
from langchain_core.messages import SystemMessage
from langgraph.checkpoint.memory import InMemorySaver
from langgraph.config import get_stream_writer
from langgraph.constants import START, END
from langgraph.graph import StateGraph

from auto_test_assistant.agents.code_generator_agent import CodeGeneratorAgentSystem
from auto_test_assistant.graph.ui_usecase_step_executor_graph import build_usecase_step_executor_agent
from auto_test_assistant.state.ui_usecase_code_generator_state import UiUseCaseCodeGeneratorState
import json
import openpyxl
from pathlib import Path


def use_case_splitting(state: UiUseCaseCodeGeneratorState):
    print(">>> use_case_splitting_node")
    writer = get_stream_writer()
    writer({"node": ">>> use_case_splitting_node"})

    uploaded_files_metadata = state.get("uploaded_files_metadata", [])
    ui_use_cases = state.get("ui_use_cases", [])

    required_columns = ["use_case_id", "func_desc", "precondition", "use_case_steps", "expect_result", "postcondition"]

    def split_test_steps(steps_str):
        if not steps_str:
            return []

        steps_list = []
        lines = str(steps_str).strip().split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            n = None
            value = line

            import re
            match = re.match(r'^(\d+)\.\s*(.*)', line)
            if match:
                n = int(match.group(1))
                value = match.group(2).strip()
            else:
                if not steps_list:
                    n = 1
                else:
                    n = steps_list[-1]["id"] + 1

            if value.endswith(';'):
                value = value[:-1]

            steps_list.append({"id": n, "value": value})

        return steps_list

    for file_meta in uploaded_files_metadata:
        file_path = file_meta.get("path")
        if not file_path:
            continue

        path = Path(file_path)
        if path.suffix.lower() not in [".xlsx", ".xls"]:
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            writer({"error": f"无法加载Excel文件 {file_path}: {e}"})
            continue

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = row
                break

            if not header_row:
                continue

            column_indices = {}
            for idx, cell_value in enumerate(header_row, start=1):
                if cell_value is None:
                    continue
                column_indices[str(cell_value).strip()] = idx

            missing_columns = [col for col in required_columns if col not in column_indices]
            if missing_columns:
                writer({"warning": f"Sheet {sheet_name} 缺少必需列: {missing_columns}"})
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                case_id = row[column_indices["use_case_id"] - 1] if column_indices["use_case_id"] <= len(row) else None
                if not case_id:
                    continue

                use_case = {"module": sheet_name}
                for col in required_columns:
                    col_idx = column_indices[col]
                    value = row[col_idx - 1] if col_idx <= len(row) else None
                    if col == "use_case_steps":
                        use_case[col] = split_test_steps(value)
                    else:
                        use_case[col] = value

                ui_use_cases.append(use_case)

    output_dir = Path("json")
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "ui_use_cases.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(ui_use_cases, f, ensure_ascii=False, indent=2)
    writer({"info": f"用例已保存到 {output_file}"})

    return {"messages": [SystemMessage(content="用例拆分完成")], "ui_use_cases": ui_use_cases}


def generate_execution_paths_from_use_cases(use_cases):
    """
    根据前置用例依赖关系生成所有可能的执行路径（拓扑排序算法）
    
    参数:
        use_cases: 用例列表，每个用例包含use_case_id和precondition字段
        
    返回:
        list[dict]: 执行路径字典列表，每个字典包含一个键值对，
                    key为目的用例ID，value为该用例的执行路径（use_case_id列表）
    """
    # 构建依赖图（正向图：前置 -> 后继）
    graph = {}
    # 反向图：后继 -> 前置，用于计算祖先
    reverse_graph = {}
    in_degree = {}
    node_to_case = {}

    # 初始化
    for case in use_cases:
        case_id = str(case.get("use_case_id", "")).strip()
        if not case_id:
            continue
        node_to_case[case_id] = case
        graph[case_id] = []
        reverse_graph[case_id] = []
        in_degree[case_id] = in_degree.get(case_id, 0)

    # 添加边
    for case in use_cases:
        case_id = str(case.get("use_case_id", "")).strip()
        if not case_id:
            continue
        precondition = case.get("precondition")
        if precondition is None:
            precondition = ""
        pre_ids = str(precondition).strip()
        if not pre_ids:
            continue
        # 按逗号分割，去除空格
        for pre_id in pre_ids.split(","):
            pre_id = pre_id.strip()
            if pre_id:
                # 确保前置节点存在于图中（可能用例列表中没有前置节点，但依赖外部）
                if pre_id not in graph:
                    # 添加一个虚拟节点
                    graph[pre_id] = []
                    reverse_graph[pre_id] = []
                    in_degree[pre_id] = 0
                    node_to_case[pre_id] = None
                graph[pre_id].append(case_id)
                reverse_graph[case_id].append(pre_id)
                in_degree[case_id] = in_degree.get(case_id, 0) + 1

    # 全局环检测：拓扑排序整个图
    global_in_degree = in_degree.copy()
    zero_degree = [node for node, deg in global_in_degree.items() if deg == 0]
    topo_order_global = []
    while zero_degree:
        u = zero_degree.pop(0)
        topo_order_global.append(u)
        for v in graph.get(u, []):
            global_in_degree[v] -= 1
            if global_in_degree[v] == 0:
                zero_degree.append(v)

    # 如果存在环，返回空字典列表
    if len(topo_order_global) != len(graph):
        # 有环，返回与用例数量相同的空字典列表
        return [{} for _ in use_cases]

    # 计算每个节点的祖先集合（传递闭包），使用反向图
    ancestors = {}

    def compute_ancestors(node, visited=None):
        if visited is None:
            visited = set()
        if node in ancestors:
            return ancestors[node]
        # 防止循环依赖
        if node in visited:
            return set()
        visited.add(node)
        result = set()
        for pre in reverse_graph.get(node, []):
            result.add(pre)
            result.update(compute_ancestors(pre, visited.copy()))
        ancestors[node] = result
        return result

    for node in graph:
        compute_ancestors(node)

    # 为每个用例构建包含所有祖先的有效执行路径
    paths = []
    for case in use_cases:
        case_id = str(case.get("use_case_id", "")).strip()
        if not case_id:
            # 用例ID为空，添加空字典
            paths.append({})
            continue

        # 获取该节点及其所有祖先的集合
        nodes_set = {case_id}
        if case_id in ancestors:
            nodes_set.update(ancestors[case_id])

        # 在子图上进行拓扑排序（使用正向图）
        subgraph = {n: [] for n in nodes_set}
        sub_in_degree = {n: 0 for n in nodes_set}
        for u in nodes_set:
            for v in graph.get(u, []):
                if v in nodes_set:
                    subgraph[u].append(v)
                    sub_in_degree[v] = sub_in_degree.get(v, 0) + 1

        # Kahn算法拓扑排序
        zero_degree = [n for n in nodes_set if sub_in_degree[n] == 0]
        topo_order = []
        while zero_degree:
            u = zero_degree.pop(0)
            topo_order.append(u)
            for v in subgraph[u]:
                sub_in_degree[v] -= 1
                if sub_in_degree[v] == 0:
                    zero_degree.append(v)

        # 检查是否所有节点都被排序（无环）——由于全局无环，子图也应无环
        if len(topo_order) != len(nodes_set):
            # 安全起见返回空字典
            paths.append({})
            continue

        # 路径即为拓扑排序结果，以字典形式存储
        paths.append({case_id: topo_order})

    return paths


def code_generator(state: UiUseCaseCodeGeneratorState):
    print(">>> code_generator_node")
    writer = get_stream_writer()
    writer({"node": ">>> code_generator_node"})
    use_cases = state.get("ui_use_cases", [])
    update_use_cases = []
    # 按前置用例，以DAG做成执行路径
    all_paths = generate_execution_paths_from_use_cases(use_cases)
    print(f"[debug] 生成的路径数量: {len(all_paths)}")
    print("all_paths", all_paths)
    for path in all_paths:
        step_executor = build_usecase_step_executor_agent()
        config = {
            "configurable": {
                "thread_id": f"step_executor_{list(path.keys())[0]}_123"
            }
        }
        use_case = next((item for item in use_cases if item.get("use_case_id") == list(path.keys())[0]), None)
        for chunk in step_executor.stream({
            "ui_use_cases": use_cases,
            "ui_use_case": use_case,
            "path": path[use_case["use_case_id"]],
            "current_use_case_idx": 0,
            "current_use_case": next((item for item in use_cases if item.get("use_case_id") == path[use_case["use_case_id"]][0]), None),
            "current_step": 0,
            "script_path": f"./scripts/use_case_{list(path.keys())[0]}.py",
            "generate_degree": 0,
            "generate_state": False
        }, config=config, stream_mode="custom", subgraphs=True):
            print(chunk)
        use_case["generate_state"] = step_executor.get_state(config=config).values.get("generate_state", False)
        use_case["script_path"] = step_executor.get_state(config=config).values.get("script_path", "")
        update_use_cases.append(use_case)

    return {"messages": [SystemMessage(content="用例全部已经转化为脚本文件")], "ui_use_cases": update_use_cases}


def code_review(state: UiUseCaseCodeGeneratorState):
    print(">>> code_review_node")
    writer = get_stream_writer()
    writer({"node": ">>> code_review_node"})
    return {"messages": [SystemMessage(content="code_review")], "type": "code_review",
            "reason": "该问题与软件测试无关"}


def build_code_generator_agent():
    """
    根据UI测试用例生成代码的graph
    """
    builder = StateGraph(UiUseCaseCodeGeneratorState)
    builder.add_node("use_case_splitting", use_case_splitting)
    builder.add_node("code_generator", code_generator)
    builder.add_node("code_review", code_review)

    builder.add_edge(START, "use_case_splitting")
    builder.add_edge("use_case_splitting", "code_generator")
    builder.add_edge("code_generator", "code_review")
    builder.add_edge("code_review", END)

    checkpoint = InMemorySaver()
    graph = builder.compile(checkpointer=checkpoint)
    return graph
